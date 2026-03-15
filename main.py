import ctypes
import io
import re
import json
import csv
from typing import List
from urllib.parse import urlparse, parse_qs

try:
    ctypes.CDLL("/usr/lib/x86_64-linux-gnu/libzbar.so.0")
except Exception:
    pass

from fastapi import FastAPI, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pyzbar import pyzbar
from PIL import Image, ImageEnhance, ImageFilter
import openpyxl

app = FastAPI(title="QR Kod Tarama Modülü")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────────
# QR OKUMA
# ─────────────────────────────────────────────

def decode_from_pil(img: Image.Image) -> list:
    """Birden fazla stratejiyle QR kod çözmeyi dener."""
    strategies = [
        img,
        img.convert("L"),
        img.convert("L").resize((img.width * 2, img.height * 2), Image.LANCZOS),
        ImageEnhance.Contrast(img.convert("L")).enhance(2.0),
        img.convert("L").filter(ImageFilter.SHARPEN),
        ImageEnhance.Sharpness(img.convert("L")).enhance(3.0),
    ]
    for s in strategies:
        try:
            codes = pyzbar.decode(s)
            if codes:
                return [c.data.decode("utf-8", errors="replace") for c in codes]
        except Exception:
            continue
    return []


def process_image_bytes(data: bytes) -> list:
    try:
        img = Image.open(io.BytesIO(data))
        return decode_from_pil(img)
    except Exception:
        return []


def process_pdf_bytes(data: bytes) -> list:
    try:
        import fitz
        doc = fitz.open(stream=data, filetype="pdf")
        results = []
        for page_num in range(len(doc)):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            qr_list = process_image_bytes(pix.tobytes("png"))
            results.extend(qr_list)
        return results
    except Exception:
        return []


# ─────────────────────────────────────────────
# VERİ AYRIŞTIRMA
# ─────────────────────────────────────────────

PLATE_TR  = re.compile(r"^\d{2}\s?[A-Z]{1,3}\s?\d{2,4}$")
PLATE_OLD = re.compile(r"^[A-Z]{2}\s?\d{4,5}$")
VIN_RE    = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$")
TC_RE     = re.compile(r"^[1-9]\d{10}$")
VERGI_RE  = re.compile(r"^\d{10}$")

EMPTY_FIELDS = {
    "ham_icerik": "",
    "dogrulama_linki": "",
    "belge_seri_no": "",
    "referans_no": "",
    "plaka": "",
    "sase_no": "",
    "ad_soyad": "",
    "tc_vergi_no": "",
    "belge_turu": "",
    "veri_tipi": "",
}


def parse_content(raw: str) -> dict:
    result = {**EMPTY_FIELDS, "ham_icerik": raw}
    s = raw.strip()
    up = s.upper().replace(" ", "")

    # URL
    if s.startswith(("http://", "https://")):
        result["veri_tipi"] = "Link"
        result["dogrulama_linki"] = s
        try:
            parsed = urlparse(s)
            params = parse_qs(parsed.query)
            for k, v in params.items():
                kl = k.lower()
                if any(x in kl for x in ["plaka", "plate"]):
                    result["plaka"] = v[0]
                elif any(x in kl for x in ["sase", "vin", "chassis"]):
                    result["sase_no"] = v[0]
                elif any(x in kl for x in ["tc", "kimlik"]):
                    result["tc_vergi_no"] = v[0]
                elif any(x in kl for x in ["vergi"]):
                    result["tc_vergi_no"] = v[0]
                elif any(x in kl for x in ["no", "number", "seri", "belge", "doc"]):
                    result["belge_seri_no"] = v[0]
                elif any(x in kl for x in ["ref", "kod", "verify", "dogrulama"]):
                    result["referans_no"] = v[0]
                elif any(x in kl for x in ["ad", "soyad", "name", "isim", "unvan"]):
                    result["ad_soyad"] = v[0]
        except Exception:
            pass
        return result

    # TC Kimlik (JSON'dan önce kontrol et)
    if TC_RE.match(s):
        result["veri_tipi"] = "TC Kimlik No"
        result["tc_vergi_no"] = s
        return result

    # Vergi No
    if VERGI_RE.match(s) and not TC_RE.match(s):
        result["veri_tipi"] = "Vergi No"
        result["tc_vergi_no"] = s
        return result

    # JSON
    try:
        data = json.loads(s)
        if not isinstance(data, dict):
            raise ValueError("not a dict")
        result["veri_tipi"] = "JSON"
        if isinstance(data, dict):
            for k, v in data.items():
                kl = k.lower()
                val = str(v)
                if any(x in kl for x in ["plaka", "plate"]):
                    result["plaka"] = val
                elif any(x in kl for x in ["sase", "vin", "chassis"]):
                    result["sase_no"] = val
                elif any(x in kl for x in ["tc", "kimlik"]):
                    result["tc_vergi_no"] = val
                elif any(x in kl for x in ["vergi"]):
                    result["tc_vergi_no"] = val
                elif any(x in kl for x in ["ad", "soyad", "name", "isim", "unvan"]):
                    result["ad_soyad"] = val
                elif any(x in kl for x in ["seri", "belge_no", "docno", "belge"]):
                    result["belge_seri_no"] = val
                elif any(x in kl for x in ["ref", "kod", "dogrulama", "verify"]):
                    result["referans_no"] = val
                elif any(x in kl for x in ["tur", "type", "belge_tur"]):
                    result["belge_turu"] = val
        return result
    except Exception:
        pass

    # VIN / Şase
    if VIN_RE.match(up):
        result["veri_tipi"] = "Şase No (VIN)"
        result["sase_no"] = up
        return result

    # Plaka
    if PLATE_TR.match(up) or PLATE_OLD.match(up):
        result["veri_tipi"] = "Plaka"
        result["plaka"] = up
        return result

    # Çok parçalı (tire veya | ile ayrılmış)
    sep = None
    if "-" in s:
        sep = "-"
    elif "|" in s:
        sep = "|"
    elif "/" in s and not s.startswith("http"):
        sep = "/"

    if sep:
        parts = s.split(sep)
        result["veri_tipi"] = "Çok Parçalı Kod"
        result["belge_seri_no"] = parts[0].strip()
        if len(parts) >= 2:
            result["referans_no"] = sep.join(p.strip() for p in parts[1:])
        return result

    # Düz alfanumerik
    if re.match(r"^[A-Za-z0-9\-_\.]+$", s):
        result["veri_tipi"] = "Seri / Referans No"
        result["belge_seri_no"] = s
        result["referans_no"] = s
        return result

    # Geri kalan
    result["veri_tipi"] = "Ham Veri"
    return result


def status_of(parsed_list: list) -> str:
    if not parsed_list:
        return "QR bulunamadı"
    types = [p.get("veri_tipi", "") for p in parsed_list]
    ham_types = {"Ham Veri", "Ayrıştırılamadı", ""}
    if all(t in ham_types for t in types):
        return "Ham veri alındı"
    if any(t in ham_types for t in types):
        return "Kısmi veri ayrıştırıldı"
    return "Veri ayrıştırıldı"


# ─────────────────────────────────────────────
# API ROTALARI
# ─────────────────────────────────────────────

@app.post("/api/scan")
async def scan_files(files: List[UploadFile] = File(...)):
    results = []

    for file in files:
        content = await file.read()
        filename = file.filename or "bilinmeyen"
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

        file_result = {
            "dosya_adi": filename,
            "durum": "",
            "qr_sayisi": 0,
            "veriler": [],
        }

        try:
            if ext == "pdf":
                qr_list = process_pdf_bytes(content)
            elif ext in ("jpg", "jpeg", "png", "bmp", "webp", "tiff", "tif"):
                qr_list = process_image_bytes(content)
            else:
                file_result["durum"] = "Desteklenmeyen format"
                results.append(file_result)
                continue

            if not qr_list:
                file_result["durum"] = "QR bulunamadı"
                results.append(file_result)
                continue

            file_result["qr_sayisi"] = len(qr_list)

            parsed_list = []
            for raw in qr_list:
                try:
                    parsed_list.append(parse_content(raw))
                except Exception:
                    parsed_list.append({
                        **EMPTY_FIELDS,
                        "ham_icerik": raw,
                        "veri_tipi": "Ayrıştırılamadı",
                    })

            file_result["durum"] = status_of(parsed_list)
            file_result["veriler"] = parsed_list

        except Exception as e:
            file_result["durum"] = f"Hata: {str(e)[:120]}"

        results.append(file_result)

    ozet = {
        "toplam": len(results),
        "qr_bulunan": sum(1 for r in results if r["qr_sayisi"] > 0),
        "qr_bulunamayan": sum(
            1 for r in results
            if r["qr_sayisi"] == 0 and r["durum"] not in ("Desteklenmeyen format",)
        ),
        "ayristirilan": sum(
            1 for r in results
            if r["durum"] in ("Veri ayrıştırıldı", "Kısmi veri ayrıştırıldı")
        ),
        "sadece_ham": sum(1 for r in results if r["durum"] == "Ham veri alındı"),
    }

    return {"ozet": ozet, "sonuclar": results}


FIELD_LABELS = {
    "ham_icerik": "Ham QR İçeriği",
    "dogrulama_linki": "Doğrulama Linki",
    "belge_seri_no": "Belge Seri No",
    "referans_no": "Referans No",
    "plaka": "Plaka",
    "sase_no": "Şase No",
    "ad_soyad": "Ad Soyad / Ünvan",
    "tc_vergi_no": "TC / Vergi No",
    "belge_turu": "Belge Türü",
    "veri_tipi": "QR Veri Tipi",
}


@app.post("/api/export/excel")
async def export_excel(request: Request):
    payload = await request.json()
    results = payload.get("sonuclar", [])
    fields = payload.get("alanlar", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QR Tarama Sonuçları"

    headers = ["Dosya Adı", "Durum", "QR Sayısı"] + [
        FIELD_LABELS.get(f, f) for f in fields if f in FIELD_LABELS
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        first = r["veriler"][0] if r.get("veriler") else {}
        row = [r["dosya_adi"], r["durum"], r["qr_sayisi"]] + [
            first.get(f, "") for f in fields if f in FIELD_LABELS
        ]
        ws.append(row)

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=qr_sonuclar.xlsx"},
    )


@app.post("/api/export/csv")
async def export_csv(request: Request):
    payload = await request.json()
    results = payload.get("sonuclar", [])
    fields = payload.get("alanlar", [])

    buf = io.StringIO()
    writer = csv.writer(buf)
    headers = ["Dosya Adı", "Durum", "QR Sayısı"] + [
        FIELD_LABELS.get(f, f) for f in fields if f in FIELD_LABELS
    ]
    writer.writerow(headers)

    for r in results:
        first = r["veriler"][0] if r.get("veriler") else {}
        row = [r["dosya_adi"], r["durum"], r["qr_sayisi"]] + [
            first.get(f, "") for f in fields if f in FIELD_LABELS
        ]
        writer.writerow(row)

    csv_bytes = "\ufeff" + buf.getvalue()  # BOM for Excel compatibility
    return StreamingResponse(
        io.BytesIO(csv_bytes.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=qr_sonuclar.csv"},
    )


# Static files (frontend) — must be last
app.mount("/", StaticFiles(directory="static", html=True), name="static")
